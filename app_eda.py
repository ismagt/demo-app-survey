# ============================================================
# Survey EDA Dashboard (Streamlit) — app_eda_v4 FIXED
#
# Fixes your two recurring errors:
# 1) ValueError: clean_cat_series expected 1 column, got 2
# 2) ValueError: Buffer has wrong number of dimensions (expected 1, got 2)
#
# Root cause: your CSV can contain DUPLICATE column names.
# In pandas, df["colname"] may return a DataFrame (2+ cols) instead of a Series.
# This breaks clean_cat_series() and pd.crosstab().
#
# This version introduces safe_get_series() and uses it everywhere.
#
# Also: includes Score as a QUESTION ROW in Tab 5 + Excel Index.
# (We create "Score (quiz) band" and we include that as a row-question AND banner if present.)
# ============================================================

import re
from io import BytesIO

import numpy as np
import pandas as pd

import streamlit as st
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go

from scipy.stats import chi2_contingency

# -----------------------------
# Optional Excel export deps
# -----------------------------
OPENPYXL_OK = True
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except Exception:
    OPENPYXL_OK = False

# -----------------------------
# Config
# -----------------------------
st.set_page_config(page_title="Survey EDA Dashboard", layout="wide")

REGION_COL = "Which region best represents the primary jurisdiction of your regulatory agency?"
AI_USE_COL = "Have you personally used AI in your professional role as a regulator?"
TRAINING_COL = "Have you or your agency participated in any formal training or workshops related to AI in the last two years?"
POSITION_COL = "Which best describes your position level within your regulatory agency?"
YEARS_COL = "How many years have you worked in a regulatory role with oversight of the gambling/gaming industry?"

SCORE_COL = "Score"
SCORE_BAND_COL = "Score (quiz) band"

ISO3_COL_CANDIDATES = ["iso3", "ISO3", "ISO_3", "country_iso3"]

ROLE_GROUP_COL = "role_group"
TENURE_GROUP_COL = "tenure_group"
REGION_GROUP_COL = "region_group"
TRAINING_BIN_COL = "training_bin"

# Columns to remove from analysis/presentation (clutter)
# IMPORTANT: Do NOT drop SCORE_COL; we use it to build SCORE_BAND_COL.
DROP_COLS = {
    "Start Date",
    "End Date",
    "Response Type",
    "Progress",
    "Duration (in seconds)",
    "Finished",
    "Recorded Date",
    "Response ID",
    "Distribution Channel",
    "User Language",
    "Q_RecaptchaScore",
    "Q_DuplicateRespondent",
    "Are you currently employed by a gambling regulatory agency, commission, authority, or equivalent public body?",
    "Do you agree to participate in this research study based on the information provided above?",
    "Source",
    "Create New Field or Choose From Dropdown...",
    "Q_UnansweredPercentage",
    "Q_UnansweredQuestions",
}

# FIXED banner columns (Tab 5 columns)
FIXED_BANNERS = [
    "Please select your age range.",
    "Please indicate your gender. - Selected Choice",
    "Please indicate your gender. - Prefer to self-describe: - Text",
    "Please indicate your ethnicity or race (select all that apply): - Selected Choice",
    "Please indicate your ethnicity or race (select all that apply): - Another race or ethnicity (please specify): - Text",
    REGION_COL,
    POSITION_COL,
    "Which division or primary area do you currently work in? - Selected Choice",
    "Which division or primary area do you currently work in? - Other (please specify): - Text",
    YEARS_COL,
    SCORE_BAND_COL,  # keep score band as a banner if present
]

# Excel styling fills
if OPENPYXL_OK:
    HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    BOLD = Font(bold=True)

# ============================================================
# Robust helpers for duplicate-column CSVs
# ============================================================

def safe_get_series(df: pd.DataFrame, colname: str) -> pd.Series:
    """
    Always return a 1D Series for a column name, even if df has duplicate column names.
    - If duplicates: take the FIRST occurrence.
    - If missing: return a Series of NaN with the right length.
    """
    if colname not in df.columns:
        return pd.Series([np.nan] * len(df), index=df.index, name=colname)

    out = df.loc[:, colname]  # can be Series or DataFrame if duplicates
    if isinstance(out, pd.DataFrame):
        out = out.iloc[:, 0]   # take first duplicate
        out.name = colname
    return out

def clean_cat_series(x) -> pd.Series:
    """
    Accept Series OR single-column DataFrame; returns clean categorical Series.
    If DataFrame with multiple cols sneaks in, we take the first.
    """
    if isinstance(x, pd.DataFrame):
        x = x.iloc[:, 0]
    s = x.copy()
    s = s.replace({np.nan: "Unknown"}).astype(str).str.strip()
    return s.replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})

def strip_strings(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    obj_cols = df2.select_dtypes(include=["object"]).columns
    for c in obj_cols:
        # Use loc[:, c] in case duplicates exist: it returns DataFrame; we handle both.
        col = df2.loc[:, c]
        if isinstance(col, pd.DataFrame):
            # strip each duplicate column
            for j in range(col.shape[1]):
                df2.iloc[:, df2.columns.get_loc(c)[j] if isinstance(df2.columns.get_loc(c), slice) else df2.columns.get_loc(c)] = (
                    col.iloc[:, j].astype(str).str.strip().replace({"": np.nan, "nan": np.nan, "None": np.nan})
                )
        else:
            df2[c] = col.astype(str).str.strip().replace({"": np.nan, "nan": np.nan, "None": np.nan})
    return df2

def coerce_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df2 = df.copy()
    for c in cols:
        if c in df2.columns:
            col = df2.loc[:, c]
            if isinstance(col, pd.DataFrame):
                # coerce each duplicate
                for j in range(col.shape[1]):
                    df2.loc[:, c].iloc[:, j] = pd.to_numeric(col.iloc[:, j], errors="coerce")
            else:
                df2[c] = pd.to_numeric(col, errors="coerce")
    return df2

def robust_barh(labels: pd.Series, values: pd.Series, title: str, xlabel: str):
    x_labels = [str(x) for x in labels.tolist()]
    y_values = values.astype(int).tolist()
    order = np.argsort(y_values)
    x_labels = [x_labels[i] for i in order]
    y_values = [y_values[i] for i in order]

    fig = plt.figure(figsize=(10, 5))
    plt.barh(x_labels, y_values)
    plt.xlabel(xlabel)
    plt.title(title)
    plt.tight_layout()
    st.pyplot(fig)

def select_filter(colname: str, df: pd.DataFrame):
    # must use safe_get_series because duplicates can exist
    if colname not in df.columns:
        return None
    s = safe_get_series(df, colname)
    vals = s.replace({np.nan: "Unknown"}).astype(str)
    options = sorted(vals.unique().tolist())
    selected = st.sidebar.multiselect(colname, options, default=options, key=f"filter_{colname}")
    return selected

# ============================================================
# Score band builder
# ============================================================

def add_score_band(df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates SCORE_BAND_COL from SCORE_COL.
    - If max >= 12 -> assume ~14-pt quiz: 0–4 / 5–9 / 10–14
    - else terciles.
    """
    df2 = df.copy()
    if SCORE_COL not in df2.columns:
        return df2

    score = pd.to_numeric(safe_get_series(df2, SCORE_COL), errors="coerce")
    df2[SCORE_COL] = score
    max_score = score.max()

    if pd.isna(max_score):
        df2[SCORE_BAND_COL] = "Unknown"
        return df2

    max_score = float(max_score)

    if max_score >= 12:
        def _band14(x):
            if pd.isna(x): return "Unknown"
            x = float(x)
            if x <= 4: return "0–4 (Low)"
            if x <= 9: return "5–9 (Medium)"
            return "10–14 (High)"
        df2[SCORE_BAND_COL] = score.map(_band14).astype(str).replace({"nan": "Unknown"}).fillna("Unknown")
        return df2

    b1 = max_score * 0.33
    b2 = max_score * 0.66

    def _band3(x):
        if pd.isna(x): return "Unknown"
        x = float(x)
        if x <= b1: return f"Low (≤{b1:.0f})"
        if x <= b2: return f"Medium ({b1:.0f}–{b2:.0f})"
        return f"High (>{b2:.0f})"

    df2[SCORE_BAND_COL] = score.map(_band3).astype(str).replace({"nan": "Unknown"}).fillna("Unknown")
    return df2

# ============================================================
# Grouping helpers
# ============================================================

def recode_role(series) -> pd.Series:
    s = clean_cat_series(series)
    leadership = {"Executive Leadership", "Director", "Commissioner", "Chair", "CEO", "President"}
    management = {"Management", "Manager", "Supervisor", "Team Lead"}
    non_mgmt = {"Non-Management", "Analyst", "Specialist", "Staff", "Individual Contributor"}

    def _map(x: str) -> str:
        if x in leadership: return "Executive Leadership"
        if x in management: return "Management"
        if x in non_mgmt: return "Non-Management"
        return "Other/Unknown"
    return s.map(_map)

def recode_training_yesno(series) -> pd.Series:
    s = clean_cat_series(series).str.lower()
    return s.map(lambda x: "Yes" if "yes" in x else ("No" if "no" in x else "Unknown"))

def recode_region_us_other(series) -> pd.Series:
    s = clean_cat_series(series).str.lower()
    return s.map(lambda x: "US" if ("united states" in x or x == "us" or "u.s." in x) else ("Other" if x != "unknown" else "Unknown"))

def recode_tenure_5yrs(series) -> pd.Series:
    s = clean_cat_series(series).str.lower()
    less = {"0-2 years", "0–2 years", "1-2 years", "3-5 years", "3–5 years", "less than 5 years", "<5 years"}
    more = {"6-10 years", "6–10 years", "10+ years", "10 years or more", "more than 5 years", ">5 years"}

    def _map(x: str) -> str:
        if x in less: return "<5 years"
        if x in more: return ">=5 years"
        return "Other/Unknown"
    return s.map(_map)

def add_group_vars(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    df2[ROLE_GROUP_COL] = recode_role(safe_get_series(df2, POSITION_COL)) if POSITION_COL in df2.columns else "Unknown"
    df2[TRAINING_BIN_COL] = recode_training_yesno(safe_get_series(df2, TRAINING_COL)) if TRAINING_COL in df2.columns else "Unknown"
    df2[REGION_GROUP_COL] = recode_region_us_other(safe_get_series(df2, REGION_COL)) if REGION_COL in df2.columns else "Unknown"
    df2[TENURE_GROUP_COL] = recode_tenure_5yrs(safe_get_series(df2, YEARS_COL)) if YEARS_COL in df2.columns else "Unknown"
    return df2

# ============================================================
# XTab helpers (counts + %)
# ============================================================

def xtab_percent_table(
    df: pd.DataFrame,
    row_var: str,
    col_var: str,
    drop_unknown_rows: bool = False,
    drop_unknown_cols: bool = False,
    sort_rows_by_total: bool = False,
) -> pd.DataFrame:
    r = clean_cat_series(safe_get_series(df, row_var))
    c = clean_cat_series(safe_get_series(df, col_var))

    tmp = pd.DataFrame({row_var: r, col_var: c}, index=df.index)

    if drop_unknown_rows:
        tmp = tmp[tmp[row_var] != "Unknown"]
    if drop_unknown_cols:
        tmp = tmp[tmp[col_var] != "Unknown"]

    ct = pd.crosstab(tmp[row_var], tmp[col_var], dropna=False)
    ct["Total"] = ct.sum(axis=1)

    col_sums = ct.sum(axis=0).replace(0, np.nan)
    pct = (ct / col_sums) * 100
    pct = pct.round(1)

    if sort_rows_by_total:
        pct = pct.sort_values("Total", ascending=False)

    return pct

def cramers_v_from_ct(ct: pd.DataFrame) -> float:
    if ct.shape[0] < 2 or ct.shape[1] < 2:
        return np.nan
    chi2, _, _, _ = chi2_contingency(ct.values)
    n = ct.to_numpy().sum()
    r, k = ct.shape
    if n <= 0:
        return np.nan
    phi2 = chi2 / n
    phi2corr = max(0, phi2 - ((k - 1) * (r - 1)) / max(1, (n - 1)))
    rcorr = r - ((r - 1) ** 2) / max(1, (n - 1))
    kcorr = k - ((k - 1) ** 2) / max(1, (n - 1))
    denom = min((kcorr - 1), (rcorr - 1))
    if denom <= 0:
        return np.nan
    return float(np.sqrt(phi2corr / denom))

# ============================================================
# Significance letters (optional) — kept from your version
# ============================================================

def _ztest_prop(p1, n1, p2, n2):
    if n1 == 0 or n2 == 0:
        return np.nan
    x1 = p1 * n1
    x2 = p2 * n2
    p = (x1 + x2) / (n1 + n2)
    se = np.sqrt(p * (1 - p) * (1/n1 + 1/n2))
    if se == 0:
        return np.nan
    z = (p1 - p2) / se
    from math import erf, sqrt
    return 2 * (1 - 0.5 * (1 + erf(abs(z) / sqrt(2))))

def sig_letter_matrix(df: pd.DataFrame, row_var: str, col_var: str, alpha95=0.05, alpha99=0.01, drop_unknown=True) -> pd.DataFrame:
    r = clean_cat_series(safe_get_series(df, row_var))
    c = clean_cat_series(safe_get_series(df, col_var))
    tmp = pd.DataFrame({row_var: r, col_var: c}, index=df.index)

    if drop_unknown:
        tmp = tmp[(tmp[row_var] != "Unknown") & (tmp[col_var] != "Unknown")]

    ct = pd.crosstab(tmp[row_var], tmp[col_var], dropna=False)
    if ct.empty:
        return pd.DataFrame()

    col_totals = ct.sum(axis=0).replace(0, np.nan)
    pct = ct.div(col_totals, axis=1)

    cols = list(ct.columns)
    letters_95 = list("abcdefghijklmnopqrstuvwxyz")
    letters_99 = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    out = pd.DataFrame("", index=ct.index, columns=ct.columns, dtype=object)

    for c_idx, c0 in enumerate(cols):
        for r0 in ct.index:
            if pd.isna(col_totals[c0]) or col_totals[c0] == 0:
                continue
            p1 = pct.loc[r0, c0]
            sig95, sig99 = [], []
            for c2_idx, c2 in enumerate(cols):
                if c2 == c0:
                    continue
                if pd.isna(col_totals[c2]) or col_totals[c2] == 0:
                    continue
                p2 = pct.loc[r0, c2]
                pval = _ztest_prop(p1, float(col_totals[c0]), p2, float(col_totals[c2]))
                if pd.isna(pval):
                    continue
                if pval < alpha99:
                    sig99.append(letters_99[c2_idx % len(letters_99)])
                elif pval < alpha95:
                    sig95.append(letters_95[c2_idx % len(letters_95)])
            out.loc[r0, c0] = "".join(sig99 + sig95)

    return out

# ============================================================
# Excel writer: XTab All Qs + Matrix
# ============================================================

def build_xtab_allqs_excel(
    df: pd.DataFrame,
    banners: list[str],
    questions: list[str],
    drop_unknown: bool = True,
    show_sig: bool = False,
    alpha95: float = 0.05,
    alpha99: float = 0.01,
) -> bytes:
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl is not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "XTab_AllQs"

    ws.append(["XTab All Questions (Column %) — Excel-style"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Notes:", "Each banner column sums to 100%. Heatmap highlights higher %."])
    ws.append(["Filters applied in Streamlit will affect this export."])
    ws.append([])

    ws_idx = wb.create_sheet("Index_Matrix")
    ws_idx.append(["Index: Questions × Banners (click to jump to the corresponding block)"])
    ws_idx["A1"].font = Font(bold=True, size=14)
    ws_idx.append([])
    start_row = ws_idx.max_row + 1
    ws_idx.cell(row=start_row, column=1, value="Question").fill = HEADER_FILL
    ws_idx.cell(row=start_row, column=1).font = BOLD

    for j, b in enumerate(banners, start=2):
        c = ws_idx.cell(row=start_row, column=j, value=b)
        c.fill = HEADER_FILL
        c.font = BOLD
        c.alignment = Alignment(wrap_text=True, vertical="top")

    anchor = {}

    for q in questions:
        ws.append([q])
        ws.cell(row=ws.max_row, column=1).fill = SECTION_FILL
        ws.cell(row=ws.max_row, column=1).font = BOLD
        ws.append([])

        for b in banners:
            anchor[(q, b)] = ws.max_row + 1

            ws.append([None, f"Banner: {b}"])
            ws.cell(row=ws.max_row, column=2).fill = HEADER_FILL
            ws.cell(row=ws.max_row, column=2).font = BOLD

            pct_tbl = xtab_percent_table(
                df=df,
                row_var=q,
                col_var=b,
                drop_unknown_rows=drop_unknown,
                drop_unknown_cols=drop_unknown,
                sort_rows_by_total=True,
            )

            banner_levels = pct_tbl.columns.tolist()
            ws.append([None, None] + banner_levels)
            hdr_row = ws.max_row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=hdr_row, column=col)
                cell.fill = HEADER_FILL
                cell.font = BOLD
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            data_start_col = 3
            first_val_row = None
            last_val_row = None

            sig_tbl = None
            if show_sig:
                sig_tbl = sig_letter_matrix(df, row_var=q, col_var=b, alpha95=alpha95, alpha99=alpha99, drop_unknown=drop_unknown)
                if not sig_tbl.empty:
                    sig_tbl = sig_tbl.reindex(columns=[c for c in banner_levels if c in sig_tbl.columns], fill_value="")

            for opt in pct_tbl.index.tolist():
                ws.append([None, opt])
                r_val = ws.max_row
                if first_val_row is None:
                    first_val_row = r_val
                last_val_row = r_val

                for j, lvl in enumerate(banner_levels):
                    v = pct_tbl.loc[opt, lvl]
                    cell = ws.cell(row=r_val, column=data_start_col + j)
                    if pd.isna(v):
                        cell.value = None
                    else:
                        cell.value = float(v) / 100.0
                        cell.number_format = "0.0%"
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                if show_sig and sig_tbl is not None and (not sig_tbl.empty) and (opt in sig_tbl.index):
                    ws.append([None, "sig"])
                    r_sig = ws.max_row
                    for j, lvl in enumerate(banner_levels):
                        cell = ws.cell(row=r_sig, column=data_start_col + j)
                        cell.value = str(sig_tbl.loc[opt, lvl]) if lvl in sig_tbl.columns else ""
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    ws.row_dimensions[r_sig].height = 14

            if first_val_row is not None and last_val_row is not None and last_val_row >= first_val_row:
                start_col = data_start_col
                end_col = data_start_col + len(banner_levels) - 1
                rng = f"{get_column_letter(start_col)}{first_val_row}:{get_column_letter(end_col)}{last_val_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                    end_type="max", end_color="C6EFCE"
                )
                ws.conditional_formatting.add(rng, rule)

            ws.append([])

        ws.append([])

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70
    for col in range(3, 60):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "A5"

    # Index hyperlinks (Score band included because it's in `questions`)
    q_start = start_row + 1
    for i, q in enumerate(questions, start=q_start):
        c = ws_idx.cell(row=i, column=1, value=q)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = SECTION_FILL
        c.font = BOLD
        for j, b in enumerate(banners, start=2):
            tgt = anchor.get((q, b))
            cell = ws_idx.cell(row=i, column=j)
            if tgt:
                cell.value = "Go"
                cell.hyperlink = f"#'XTab_AllQs'!A{tgt}"
                cell.style = "Hyperlink"
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_idx.column_dimensions["A"].width = 70
    for j in range(2, 2 + len(banners)):
        ws_idx.column_dimensions[get_column_letter(j)].width = 22
    ws_idx.freeze_panes = ws_idx["B" + str(q_start)].coordinate

    # Cramer's V matrix sheet
    ws_v = wb.create_sheet("QxBanner_Assoc_V")
    ws_v.append(["Question × Banner association (Cramer's V)"])
    ws_v["A1"].font = Font(bold=True, size=14)
    ws_v.append(["Interpretation:", "0=none, 1=perfect association (categorical)."])
    ws_v.append([])

    ws_v.append(["Question"] + banners)
    for col in range(1, 2 + len(banners)):
        c = ws_v.cell(row=ws_v.max_row, column=col)
        c.fill = HEADER_FILL
        c.font = BOLD
        c.alignment = Alignment(wrap_text=True, vertical="top")

    first_data_row = ws_v.max_row + 1
    for q in questions:
        row = [q]
        for b in banners:
            rq = clean_cat_series(safe_get_series(df, q))
            rb = clean_cat_series(safe_get_series(df, b))
            tmp = pd.DataFrame({q: rq, b: rb}, index=df.index)
            if drop_unknown:
                tmp = tmp[(tmp[q] != "Unknown") & (tmp[b] != "Unknown")]
            ct = pd.crosstab(tmp[q], tmp[b], dropna=False)
            row.append(cramers_v_from_ct(ct))
        ws_v.append(row)

    last_data_row = ws_v.max_row
    if last_data_row >= first_data_row:
        start_col = 2
        end_col = 1 + len(banners)
        rng = f"{get_column_letter(start_col)}{first_data_row}:{get_column_letter(end_col)}{last_data_row}"
        rule = ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE"
        )
        ws_v.conditional_formatting.add(rng, rule)

    ws_v.column_dimensions["A"].width = 70
    for j in range(2, 2 + len(banners)):
        ws_v.column_dimensions[get_column_letter(j)].width = 18
    ws_v.freeze_panes = "B5"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ============================================================
# App UI
# ============================================================

st.title("Survey EDA Dashboard (Interactive)")

st.sidebar.header("Data Source")
uploaded = st.sidebar.file_uploader("Upload CSV (required)", type=["csv"], key="csv_uploader")
if uploaded is None:
    st.warning("⬅️ Please upload a CSV using the sidebar to start the dashboard.")
    st.stop()

df = pd.read_csv(uploaded)
df = strip_strings(df)

# Drop clutter columns
drop_present = [c for c in DROP_COLS if c in df.columns]
if drop_present:
    df = df.drop(columns=drop_present)

# Coerce numeric (including Score)
df = coerce_numeric(df, [
    "Duration (in seconds)", "Progress", "Q_RecaptchaScore", SCORE_COL,
    "Q_UnansweredPercentage", "Q_UnansweredQuestions"
])

# Fix known typo(s) in region (use safe_get_series)
if REGION_COL in df.columns:
    reg = safe_get_series(df, REGION_COL)
    reg = reg.replace({np.nan: "Unknown"}).astype(str).str.strip().replace({"Asia-Paciifc": "Asia-Pacific", "nan": "Unknown", "": "Unknown"})
    # assign back safely
    df[REGION_COL] = reg

# Build Score band (so Score appears as a row-question via band)
df = add_score_band(df)

# -----------------------------
# Filters
# -----------------------------
st.sidebar.header("Filters")
sel_region = select_filter(REGION_COL, df)
sel_ai = select_filter(AI_USE_COL, df)
sel_training = select_filter(TRAINING_COL, df)
sel_position = select_filter(POSITION_COL, df)
sel_years = select_filter(YEARS_COL, df)

df_f = df.copy()
for col, sel in [
    (REGION_COL, sel_region),
    (AI_USE_COL, sel_ai),
    (TRAINING_COL, sel_training),
    (POSITION_COL, sel_position),
    (YEARS_COL, sel_years),
]:
    if sel is not None and col in df_f.columns:
        s = safe_get_series(df_f, col).replace({np.nan: "Unknown"}).astype(str)
        df_f = df_f[s.isin(sel)]

df_f = add_group_vars(df_f)

st.caption(f"Rows after filters: {len(df_f)} / {len(df)}")

with st.expander("Preview data"):
    st.dataframe(df_f.head(50), use_container_width=True)

# -----------------------------
# Tabs
# -----------------------------
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "Univariate", "Bivariate", "Correlations", "Geo / Map",
    "XTab All Qs (Excel-style)", "Visual XTab Matrix"
])

# -----------------------------
# Tab 1: Univariate
# -----------------------------
with tab1:
    st.subheader("Univariate distributions")

    num_cols = df_f.select_dtypes(include=["number"]).columns.tolist()
    if len(num_cols) == 0:
        st.info("No numeric columns detected.")
    else:
        col = st.selectbox("Numeric variable (histogram)", num_cols, index=0, key="univ_num_col")
        bins = st.slider("Bins", 5, 60, 20, key="univ_bins")

        fig = plt.figure(figsize=(8, 4))
        x = pd.to_numeric(safe_get_series(df_f, col), errors="coerce").dropna()
        plt.hist(x, bins=bins)
        plt.title(f"Histogram: {col}")
        plt.xlabel(col)
        plt.ylabel("Count")
        plt.tight_layout()
        st.pyplot(fig)

        fig2 = plt.figure(figsize=(6, 2.5))
        plt.boxplot(x, vert=False)
        plt.title(f"Boxplot: {col}")
        plt.tight_layout()
        st.pyplot(fig2)

    st.subheader("Categorical frequency")
    cat_cols = df_f.select_dtypes(include=["object"]).columns.tolist()
    if len(cat_cols) == 0:
        st.info("No categorical columns detected.")
    else:
        cat = st.selectbox("Categorical variable", cat_cols, index=0, key="univ_cat_col")
        s = safe_get_series(df_f, cat).replace({np.nan: "Unknown"}).astype(str)
        vc = s.value_counts(dropna=False).reset_index()
        vc.columns = ["Category", "Count"]
        st.dataframe(vc, use_container_width=True)
        robust_barh(vc["Category"], vc["Count"], title=f"Counts: {cat}", xlabel="Count")

# -----------------------------
# Tab 2: Bivariate
# -----------------------------
with tab2:
    st.subheader("Bivariate exploration")

    num_cols = df_f.select_dtypes(include=["number"]).columns.tolist()
    cat_cols = df_f.select_dtypes(include=["object"]).columns.tolist()

    if len(num_cols) == 0 or len(cat_cols) == 0:
        st.info("Need at least one numeric and one categorical column.")
    else:
        y = st.selectbox("Numeric Y", num_cols, index=0, key="biv_y")
        x = st.selectbox("Categorical X", cat_cols, index=0, key="biv_x")

        tmp = pd.DataFrame({
            x: safe_get_series(df_f, x).replace({np.nan: "Unknown"}).astype(str),
            y: pd.to_numeric(safe_get_series(df_f, y), errors="coerce"),
        }, index=df_f.index).dropna(subset=[y])

        top_k = st.slider("Max categories to show (top-k)", 3, 25, 12, key="biv_topk")
        topcats = tmp[x].value_counts().head(top_k).index
        tmp2 = tmp[tmp[x].isin(topcats)]

        fig = plt.figure(figsize=(10, 4))
        cats_sorted = tmp2[x].value_counts().index.tolist()

        data = []
        labels = []
        for c in cats_sorted:
            arr = tmp2.loc[tmp2[x] == c, y].dropna().values
            if len(arr) > 0:
                data.append(arr)
                labels.append(c)

        if len(data) < 2:
            st.info("Not enough non-null numeric data to build a boxplot for the selected groups.")
        else:
            plt.boxplot(data, labels=labels, vert=True)
            plt.xticks(rotation=30, ha="right")
            plt.title(f"{y} by {x} (top {top_k})")
            plt.tight_layout()
            st.pyplot(fig)

# -----------------------------
# Tab 3: Correlations (Spearman) — lightweight
# -----------------------------
with tab3:
    st.subheader("Correlation Matrix (Spearman)")

    likert_cols = []
    for c in df_f.columns:
        s = pd.to_numeric(safe_get_series(df_f, c), errors="coerce")
        if s.notna().sum() == 0:
            continue
        ok = s.between(1, 5).mean()
        if ok >= 0.70:
            likert_cols.append(c)

    st.write(f"Detected Likert/ordinal (1–5) columns: {len(likert_cols)}")

    min_nonnull = st.slider("Min non-null responses per variable", 5, 100, 10, key="corr_min_nonnull")
    cols_use = [c for c in likert_cols if pd.to_numeric(safe_get_series(df_f, c), errors="coerce").notna().sum() >= min_nonnull]

    if len(cols_use) < 2:
        st.info("Not enough Likert/ordinal columns after thresholds.")
    else:
        X = pd.DataFrame({c: pd.to_numeric(safe_get_series(df_f, c), errors="coerce") for c in cols_use})
        r_mat = X.corr(method="spearman")
        st.dataframe(r_mat.round(3), use_container_width=True)

# -----------------------------
# Tab 4: Geo / Map
# -----------------------------
with tab4:
    st.subheader("Geographic view (no pycountry)")

    iso3_col = next((c for c in ISO3_COL_CANDIDATES if c in df_f.columns), None)
    if iso3_col:
        tmp = safe_get_series(df_f, iso3_col).replace({np.nan: "Unknown"}).astype(str).str.strip()
        tmp = tmp[tmp.str.len() == 3]
        counts = tmp.value_counts().reset_index()
        counts.columns = ["iso3", "respondents"]
        fig = px.choropleth(counts, locations="iso3", color="respondents", title="Respondents by Country (ISO3)")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No ISO3 column detected. Add ISO3 codes to enable a world map.")

# -----------------------------
# Tab 5: XTab All Qs (Excel-style) — FIXED banners
# Includes Score band as a ROW QUESTION (and also as banner if present)
# -----------------------------
with tab5:
    st.subheader("Excel-style XTab (All Questions) — Preview + Export")
    st.caption("This tab excludes metadata columns automatically and uses your FIXED demographic banners as columns.")

    df_work = df_f.drop(columns=[c for c in DROP_COLS if c in df_f.columns], errors="ignore")

    # Banners: fixed list (present only)
    banners = [c for c in FIXED_BANNERS if c in df_work.columns]
    if len(banners) == 0:
        st.error("None of the FIXED_BANNERS columns were found in the uploaded CSV. Check column names.")
        st.stop()

    st.markdown("### Fixed banner columns (used as Excel columns)")
    st.write(banners)

    # Questions: all other categorical columns (not in banners, not internal)
    exclude_internal = {ROLE_GROUP_COL, TENURE_GROUP_COL, REGION_GROUP_COL, TRAINING_BIN_COL}
    cat_cols = df_work.select_dtypes(include=["object"]).columns.tolist()
    cat_cols = [c for c in cat_cols if c not in exclude_internal]

    questions_all = [c for c in cat_cols if c not in set(banners)]

    # FORCE: include Score band as a ROW QUESTION (even if it is also a banner)
    if SCORE_BAND_COL in df_work.columns:
        questions_all = [q for q in questions_all if q != SCORE_BAND_COL] + [SCORE_BAND_COL]

    if len(questions_all) == 0:
        st.warning("No remaining categorical question columns after applying fixed banners.")
        st.stop()

    # Preview controls
    st.markdown("### Preview (Excel-like table: column % + heatmap)")
    c1, c2, c3, c4 = st.columns([2, 2, 1, 1])
    with c1:
        preview_q = st.selectbox("Question (rows)", options=questions_all, key="xtab_preview_q")
    with c2:
        preview_banner = st.selectbox("Banner (columns)", options=banners, key="xtab_preview_b")
    with c3:
        drop_unknown = st.checkbox("Drop Unknown", value=True, key="xtab_drop_unknown")
    with c4:
        show_sig = st.checkbox("Sig letters", value=True, key="xtab_show_sig")

    alpha95 = st.slider("Alpha (95% letters)", 0.001, 0.10, 0.05, step=0.001, key="xtab_alpha95")
    alpha99 = st.slider("Alpha (99% letters)", 0.001, 0.10, 0.01, step=0.001, key="xtab_alpha99")

    pct = xtab_percent_table(
        df_work, row_var=preview_q, col_var=preview_banner,
        drop_unknown_rows=drop_unknown, drop_unknown_cols=drop_unknown, sort_rows_by_total=True
    )
    st.dataframe(
        pct.style.format(lambda v: "" if pd.isna(v) else f"{v:.1f}%").background_gradient(axis=None),
        use_container_width=True
    )

    if show_sig:
        sig = sig_letter_matrix(
            df_work, row_var=preview_q, col_var=preview_banner,
            alpha95=alpha95, alpha99=alpha99, drop_unknown=drop_unknown
        )
        if not sig.empty:
            st.caption("Significance letters within each banner column: a/b/c = 95%, A/B/C = 99% (approx z-test).")
            st.dataframe(sig, use_container_width=True)

    st.divider()

    # Matrix view: Questions x Banners (Cramer's V) — robust to duplicate columns
    st.markdown("### Question × Banner matrix (association)")
    st.caption("Cramer's V computed from the contingency table (after dropping Unknown if selected).")

    v_rows = []
    for q in questions_all:
        row = {"Question": q}
        rq = clean_cat_series(safe_get_series(df_work, q))
        for b in banners:
            rb = clean_cat_series(safe_get_series(df_work, b))
            tmp = pd.DataFrame({q: rq, b: rb}, index=df_work.index)
            if drop_unknown:
                tmp = tmp[(tmp[q] != "Unknown") & (tmp[b] != "Unknown")]
            ct = pd.crosstab(tmp[q], tmp[b], dropna=False)
            row[b] = cramers_v_from_ct(ct)
        v_rows.append(row)

    v_df = pd.DataFrame(v_rows).set_index("Question")
    st.dataframe(
        v_df.style.format("{:.3f}").background_gradient(axis=None),
        use_container_width=True,
        height=420
    )

    st.divider()

    # Excel export
    st.markdown("### Build full Excel report (All Questions × FIXED Banners)")
    if not OPENPYXL_OK:
        st.error("openpyxl is not installed in this environment. Add it to requirements.txt: openpyxl")
    else:
        if st.button("Build XTab All Qs Excel", key="build_xtab_allqs_btn"):
            xls_bytes = build_xtab_allqs_excel(
                df=df_work,
                banners=banners,
                questions=questions_all,  # includes Score band row
                drop_unknown=drop_unknown,
                show_sig=show_sig,
                alpha95=alpha95,
                alpha99=alpha99
            )
            st.download_button(
                "Download XTab All Qs (Excel heatmap + matrix)",
                data=xls_bytes,
                file_name="XTab_AllQs_ExcelStyle.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_xtab_allqs"
            )

# -----------------------------
# Tab 6: Visual XTab Matrix (Plotly)
# -----------------------------
with tab6:
    st.subheader("Visual XTab Matrix (colored % heatmaps)")

    cat_cols = df_f.select_dtypes(include=["object"]).columns.tolist()
    exclude = {ROLE_GROUP_COL, TENURE_GROUP_COL, REGION_GROUP_COL, TRAINING_BIN_COL}
    cat_cols = [c for c in cat_cols if c not in exclude]

    if len(cat_cols) < 2:
        st.info("Need at least two categorical columns.")
    else:
        default_banners = [c for c in [REGION_COL, POSITION_COL, AI_USE_COL, TRAINING_COL, YEARS_COL] if c in cat_cols]
        banners_sel = st.multiselect(
            "Banners (grouping variables)",
            options=cat_cols,
            default=default_banners[:3] if default_banners else cat_cols[:2],
            key="xtab_matrix_banners"
        )

        candidate_questions = [c for c in cat_cols if c not in set(banners_sel)]
        questions_sel = st.multiselect(
            "Questions (to cross vs each banner)",
            options=candidate_questions,
            default=candidate_questions[:3] if len(candidate_questions) >= 3 else candidate_questions,
            key="xtab_matrix_questions"
        )

        drop_unknown2 = st.checkbox("Drop Unknown (recommended)", value=True, key="xtab_matrix_drop_unknown")
        max_rows = st.slider("Max response options (rows) per heatmap", 5, 60, 25, key="xtab_matrix_max_rows")
        max_cols = st.slider("Max banner categories (cols) per heatmap", 3, 30, 12, key="xtab_matrix_max_cols")

        if len(banners_sel) == 0 or len(questions_sel) == 0:
            st.info("Select at least 1 banner and 1 question.")
        else:
            for q in questions_sel:
                st.markdown(f"## {q}")
                for b in banners_sel:
                    try:
                        pct = xtab_percent_table(
                            df_f, row_var=q, col_var=b,
                            drop_unknown_rows=drop_unknown2, drop_unknown_cols=drop_unknown2,
                            sort_rows_by_total=True
                        )

                        pct2 = pct.copy()
                        if pct2.shape[1] > max_cols:
                            col_order = pct2.sum(axis=0).sort_values(ascending=False).index[:max_cols]
                            pct2 = pct2[col_order]
                        if pct2.shape[0] > max_rows:
                            row_order = pct2.sum(axis=1).sort_values(ascending=False).index[:max_rows]
                            pct2 = pct2.loc[row_order]

                        z = pct2.values
                        text = np.vectorize(lambda x: "" if np.isnan(x) else f"{x:.1f}%")(z)

                        fig = go.Figure(
                            data=go.Heatmap(
                                z=z,
                                x=[str(c) for c in pct2.columns],
                                y=[str(r) for r in pct2.index],
                                text=text,
                                texttemplate="%{text}",
                                hovertemplate="Row: %{y}<br>Col: %{x}<br>%: %{z:.1f}<extra></extra>",
                                colorbar={"title": "%"},
                            )
                        )
                        fig.update_layout(
                            title=f"{q} × {b} (column %)",
                            height=max(420, 28 * (pct2.shape[0] + 6)),
                            margin=dict(l=10, r=10, t=50, b=10),
                        )
                        st.plotly_chart(fig, use_container_width=True)

                    except Exception as e:
                        st.warning(f"Could not plot {q} × {b}: {e}")
